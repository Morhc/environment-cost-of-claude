"""Extract EPA AVERT regional marginal CO2 emission rates for the AVERT regions that
host Anthropic-linked compute, and compare them with the eGRID annual averages used
elsewhere in this project.

Source: EPA, "Emission Rates from AVERT" (April 2024), data year 2023.
  https://www.epa.gov/system/files/documents/2024-04/avert_emission_rates_04-11-24_0.xlsx
Sheet "2023", "Regional Emission Rates" block, column "Uniform EE" — the flat, all-hours
displacement profile, i.e. the one that matches a 24/7 data-center load.

Prints a JSON block for pasting into data/sourced_data.json; does not write it, so the
hand-curated formatting of that file is preserved.

Requires openpyxl (not a dependency of the figure scripts):
    python3 -m venv .venv && .venv/bin/pip install openpyxl
    .venv/bin/python extract_avert.py
Run from the project root.
"""
import json
import os
import urllib.request

URL = ("https://www.epa.gov/system/files/documents/2024-04/"
       "avert_emission_rates_04-11-24_0.xlsx")
XLSX = "data/avert_emission_rates_2023.xlsx"
LB_PER_MWH_TO_G_PER_KWH = 453.59237 / 1000

# AVERT region -> the Anthropic-linked site it covers, and the eGRID subregion average
# the main report uses for that site. The two geographies are NOT coextensive; see the
# note emitted at the bottom.
SITES = {
    "Mid-Atlantic": ("New Carlisle, IN (Project Rainier)", "RFCW", 413),
    "Texas":        ("Barber Lake / Abernathy, TX",        "ERCT", 333),
    "New York":     ("Lake Mariner, Barker, NY",           "NYUP", 110),
    "Midwest":      ("Hawesville, KY (MISO/Big Rivers)",   None,   None),
    "Tennessee":    ("Memphis, TN (grid-served portion)",  None,   None),
}
US_AVG_EGRID = 348  # gCO2/kWh, eGRID 2023 Rev 2

if not os.path.exists(XLSX):
    os.makedirs("data", exist_ok=True)
    print(f"downloading {URL}")
    urllib.request.urlretrieve(URL, XLSX)

import openpyxl  # noqa: E402  (imported after the download so the error is clearer)

ws = openpyxl.load_workbook(XLSX, data_only=True)["2023"]

# Locate the "Uniform EE" column and the regional CO2 block by header text rather than
# by fixed offsets, so a re-released workbook fails loudly instead of silently shifting.
header_row = next(i for i, r in enumerate(ws.iter_rows(values_only=True), 1)
                  if r[0] and str(r[0]).startswith("Avoided CO2 Rate (lb/MWh)"))
cols = [str(v) if v is not None else "" for v in
        next(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))]
col = cols.index("Uniform EE")

rates = {}
for row in ws.iter_rows(min_row=header_row + 2, max_row=header_row + 16, values_only=True):
    if row[0] and isinstance(row[col], (int, float)):
        rates[str(row[0]).strip()] = float(row[col])

# National weighted average, from the "National Emission Rates" block at the top.
nat_cols = [str(v) if v is not None else "" for v in
            next(ws.iter_rows(min_row=6, max_row=6, values_only=True))]
national = float(next(r for r in ws.iter_rows(min_row=7, max_row=7, values_only=True))
                 [nat_cols.index("Uniform EE")])

print(f"{'AVERT region':<16} {'lb CO2/MWh':>11} {'g CO2/kWh':>10}   vs eGRID avg")
print("-" * 68)
out = {}
for region, (site, egrid_sub, egrid_val) in SITES.items():
    lb = rates[region]
    g = lb * LB_PER_MWH_TO_G_PER_KWH
    out[region] = {"lb_co2_per_mwh": round(lb, 1), "gco2_per_kwh": round(g, 1),
                   "site": site, "egrid_subregion": egrid_sub,
                   "egrid_gco2_per_kwh": egrid_val}
    if egrid_val:
        ratio = g / egrid_val
        out[region]["marginal_over_average"] = round(ratio, 2)
        cmp = f"{egrid_sub} {egrid_val} -> {ratio:.2f}x"
    else:
        cmp = "(no subregion average assigned in report)"
    print(f"{region:<16} {lb:>11.1f} {g:>10.1f}   {cmp}")
    print(f"{'':<16} {site}")

nat_g = national * LB_PER_MWH_TO_G_PER_KWH
print("-" * 68)
print(f"{'US national':<16} {national:>11.1f} {nat_g:>10.1f}   "
      f"eGRID US avg {US_AVG_EGRID} -> {nat_g / US_AVG_EGRID:.2f}x")

print("\n--- JSON block for data/sourced_data.json ---")
print(json.dumps({"avert_2023": {
    "source": "EPA, 'Emission Rates from AVERT' (April 2024 release), data year 2023",
    "url": URL,
    "method": ("Short-run marginal (avoided) CO2 rates from AVERT's hourly dispatch model, "
               "assuming 0.5% displacement of existing regional demand. 'Uniform EE' column = "
               "flat all-hours load profile, matching a 24/7 data-center load. Rates are "
               "T&D-loss adjusted (retail-level) per the workbook README."),
    "units": "lb CO2/MWh as published; gCO2/kWh derived at 453.59237 g/lb",
    "caveats": [
        "AVERT regions are aggregations of balancing authorities and are NOT coextensive with "
        "eGRID subregions; marginal/average ratios below cross geographic boundaries.",
        "Indiana is split 21% Mid-Atlantic / 79% Midwest; New Carlisle is served by Indiana "
        "Michigan Power, a PJM member, placing it in the Mid-Atlantic region.",
        "AVERT's New York region spans all of NYISO, including gas-heavy downstate, while "
        "eGRID NYUP is upstate only — the NY ratio is inflated by this mismatch.",
        "EPA states AVERT 'is a marginal emissions assessment tool and not a tool for emissions "
        "accounting' and cautions against its use in corporate GHG reporting.",
        "T&D adjustment makes these modestly high for a transmission-connected campus.",
    ],
    "co2_gco2_per_kwh": {**{k: v["gco2_per_kwh"] for k, v in out.items()},
                         "US_national": round(nat_g, 1)},
    "detail": out,
}}, indent=2))
