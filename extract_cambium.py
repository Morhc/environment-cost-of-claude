"""Extract NREL Cambium 2023 long-run marginal emission rates (LRMER) for the GEA regions
hosting Anthropic-linked compute.

Why this and not only AVERT: AVERT publishes SHORT-run marginal rates and its own documentation
says they "should not be used to examine the emission impacts of changes that extend more than 5
years into the future." Anthropic's Hawesville lease is 20 years and the campuses are multi-decade
assets, so the horizon-appropriate convention is the LONG-run marginal rate. Gagnon & Cole (PNAS
2022) make exactly this argument, and Cambium is the dataset they point to.

Source: NREL, "Long-Run Marginal Emission Rates for Electricity - Workbooks for 2023 Cambium Data"
  https://data.openei.org/submissions/8279
Workbook defaults, used as published (see the "Levelized LRMER" sheet):
  start year 2025 | evaluation period 20 years | real discount rate 3% | scenario Mid-case
  GWP 100-year AR6 | location End-use (so distribution losses are included)
  units kg CO2 per MWh at the point of end-use == gCO2/kWh

Prints a JSON block for pasting into data/sourced_data.json.

Requires openpyxl:  .venv/bin/python extract_cambium.py    (run from the project root)
"""
import json
import os
import urllib.request

URL = ("https://data.nlr.gov/system/files/230/"
       "1707947178-Cambium23_LRMER_GEAregions_0.xlsx")
XLSX = "data/cambium23_lrmer_gearegions.xlsx"   # 17 MB — gitignored, refetched on demand

# GEA region -> the Anthropic-linked site it serves, and the eGRID subregion the report uses.
SITES = {
    "PJM_West":     ("New Carlisle, IN (Project Rainier)", "RFCW", 413),
    "ERCOT":        ("Barber Lake / Abernathy, TX",        "ERCT", 333),
    "NYISO":        ("Lake Mariner, Barker, NY",           "NYUP", 110),
    "MISO_Central": ("Hawesville, KY",                     None,   None),
}

if not os.path.exists(XLSX):
    os.makedirs("data", exist_ok=True)
    print(f"downloading {URL} (~17 MB)")
    urllib.request.urlretrieve(URL, XLSX)

import openpyxl  # noqa: E402

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

# Confirm the levelization parameters are the published defaults rather than someone's edits.
lev = {}
for row in wb["Levelized LRMER"].iter_rows(min_row=6, max_row=15, values_only=True):
    if row[1] is not None and row[3] is not None:
        lev[str(row[1]).strip()] = row[3]
expected = {"Start year": 2025, "Evaluation period (years)": 20, "Discount rate (real)": 0.03,
            "Scenario": "Mid-case", "Emission": "CO2", "Emission stage": "Combustion"}
for k, v in expected.items():
    got = lev.get(k)
    assert got == v, f"workbook default changed: {k} is {got!r}, expected {v!r}"
print("levelization parameters (as published):")
for k, v in lev.items():
    print(f"  {k:28s} {v}")

# "Data - Annual": col 1 = region, col 2 = combustion CO2, col 13 = combined (incl. precombustion)
# CO2e, both already levelized on the parameters above.
rates = {}
for row in wb["Data - Annual"].iter_rows(min_row=6, max_row=30, values_only=True):
    if row[1] and isinstance(row[2], (int, float)):
        rates[str(row[1]).strip()] = (float(row[2]), float(row[13]))

print(f"\n{'GEA region':<14} {'LRMER CO2':>10} {'CO2e':>8}   vs eGRID annual average")
print("-" * 74)
out = {}
for region, (site, sub, avg) in SITES.items():
    co2, co2e = rates[region]
    out[region] = {"lrmer_co2_gco2_per_kwh": round(co2, 1),
                   "lrmer_co2e_gco2e_per_kwh": round(co2e, 1),
                   "site": site, "egrid_subregion": sub, "egrid_gco2_per_kwh": avg}
    if avg:
        out[region]["lrmer_over_average"] = round(co2 / avg, 2)
        cmp = f"{sub} {avg} -> {co2/avg:.2f}x"
    else:
        cmp = "(no subregion average assigned in report)"
    print(f"{region:<14} {co2:>10.1f} {co2e:>8.1f}   {cmp}")
    print(f"{'':<14} {site}")

sited = {r: out[r]["lrmer_co2_gco2_per_kwh"] for r in ("PJM_West", "ERCOT", "NYISO")}
lo, hi = min(sited.values()), max(sited.values())
print("-" * 74)
print(f"spread across the three sited regions: {lo:.0f}-{hi:.0f} gCO2/kWh = {hi/lo:.2f}x")
print(f"  (the same three on eGRID annual averages: 110-413 = 3.75x)")
print(f"  cleanest on this basis is {min(sited, key=sited.get)}, which is NOT the cleanest "
      f"on an annual-average basis")

print("\n--- JSON block for data/sourced_data.json ---")
print(json.dumps({"cambium_2023_lrmer": {
    "source": "NREL, 'Long-Run Marginal Emission Rates for Electricity — Workbooks for 2023 "
              "Cambium Data' (released Feb 2024), sheet 'Data - Annual'",
    "url": "https://data.openei.org/submissions/8279",
    "method": ("Long-run marginal emission rate: the rate of emissions induced or avoided by a "
               "sustained change in demand, accounting for the build and retirement of capital "
               "assets as well as dispatch. Levelized on the workbook's published defaults."),
    "levelization": {k: v for k, v in lev.items()},
    "units": "kg CO2/MWh at point of end-use == gCO2/kWh; distribution losses included",
    "extracted_by": "extract_cambium.py",
    "gco2_per_kwh": {r: v["lrmer_co2_gco2_per_kwh"] for r, v in out.items()},
    "gco2e_per_kwh_incl_precombustion": {r: v["lrmer_co2e_gco2e_per_kwh"] for r, v in out.items()},
    "site_mapping": {r: v["site"] for r, v in out.items()},
    "spread_across_sited_regions": round(hi / lo, 2),
    "caveats": [
        "LRMER is the horizon-appropriate convention here: NREL advises applying it to "
        "interventions persisting five years or longer, and AVERT's short-run rates explicitly "
        "should not be used beyond a 5-year horizon. Anthropic's Hawesville lease is 20 years.",
        "Values are scenario-dependent (Mid-case shown); the workbook carries 8 scenarios "
        "including high/low gas price and 95%/100% decarbonization.",
        "GEA regions are not coextensive with eGRID subregions, so LRMER-vs-average ratios cross "
        "boundaries. The within-Cambium comparison across the three sites does not.",
        "Cambium NYISO spans all of New York, not upstate alone.",
        "'End-use' location means distribution losses are included, modestly high for a "
        "transmission-connected campus.",
    ],
}}, indent=2))
